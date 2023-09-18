from simple_spider.pudong_school.spider.mongo_client import get_mongo_client
from openpyxl import load_workbook


def load_excel():
    workbook = load_workbook(
        filename='../data/学校情况填充表2.xlsx')
    worksheet = workbook['填充后']
    index = 1
    for row in range(2, worksheet.max_row + 1):
        # 读取某一列的值
        column_value = worksheet['A' + str(row)].value  # 假设要读取的列为A列
        print("获取学校【{}】".format(column_value))
        print("当前索引位置【{}】".format(index))
        index = index + 1
        # if school_location is not None:
        #     location_json = school_location.get("location_json")
        #     formatted_address = location_json['geocodes'][0]['formatted_address']
        #     location = location_json['geocodes'][0]['location']
        #     location_split = location.split(",")
        #     print("获取学校查询地址【{}】".format(formatted_address))
        #     # 根据读取的值填充另一列
        #     worksheet['M' + str(row)].value = formatted_address  # 假设要填充的列为B列，这里将读取的值乘以2填充
        #     worksheet['C' + str(row)].value = location_split[0]  # 假设要填充的列为B列，这里将读取的值乘以2填充
        #     worksheet['D' + str(row)].value = location_split[1]  # 假设要填充的列为B列，这里将读取的值乘以2填充

    # 保存修改后的Excel文件
    # workbook.save('../data/学校情况填充表3.xlsx')
    workbook.close()

def load_excel_elementary_school():
    workbook = load_workbook(
        filename='../data/浦东小学招生信息公开.xls.xlsx')
    worksheet = workbook['表一']
    index = 1
    for row in range(2, worksheet.max_row + 1):
        # 读取某一列的值
        column_value = worksheet['A' + str(row)].value  # 假设要读取的列为A列
        print("获取学校【{}】".format(column_value))
        print("当前索引位置【{}】".format(index))
        index = index + 1
        # if school_location is not None:
        #     location_json = school_location.get("location_json")
        #     formatted_address = location_json['geocodes'][0]['formatted_address']
        #     location = location_json['geocodes'][0]['location']
        #     location_split = location.split(",")
        #     print("获取学校查询地址【{}】".format(formatted_address))
        #     # 根据读取的值填充另一列
        #     worksheet['M' + str(row)].value = formatted_address  # 假设要填充的列为B列，这里将读取的值乘以2填充
        #     worksheet['C' + str(row)].value = location_split[0]  # 假设要填充的列为B列，这里将读取的值乘以2填充
        #     worksheet['D' + str(row)].value = location_split[1]  # 假设要填充的列为B列，这里将读取的值乘以2填充

    # 保存修改后的Excel文件
    # workbook.save('../data/学校情况填充表3.xlsx')
    workbook.close()


if __name__ == '__main__':
    """
    1、读取位置信息
    2、解析出excel
    """
    load_excel()
