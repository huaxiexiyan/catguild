import re

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook


def run():
    workbook = load_workbook(filename='../data/目标位置5.xlsx')
    source_worksheet = workbook['位置']

    # 创建两个新的工作簿用于存储包含和不包含的记录
    included_workbook = Workbook()
    included_worksheet = included_workbook.active
    included_worksheet.title = '位置'

    excluded_workbook = Workbook()
    excluded_worksheet = excluded_workbook.active
    excluded_worksheet.title = '位置'

    error_workbook = Workbook()
    error_worksheet = error_workbook.active
    error_worksheet.title = '位置'

    # 复制原始数据的标题行到新工作簿
    included_worksheet.append([cell.value for cell in source_worksheet[1]])
    excluded_worksheet.append([cell.value for cell in source_worksheet[1]])
    error_worksheet.append([cell.value for cell in source_worksheet[1]])

    for row in range(2, source_worksheet.max_row + 1):
        # 读取某一列的值
        url = source_worksheet['H' + str(row)].value  # 假设要读取的列为A列
        # url = source_worksheet['I' + str(row)].value  # 假设要读取的列为A列
        location = source_worksheet['E' + str(row)].value  # 假设要读取的列为A列
        # location_rep = source_worksheet['C' + str(row)].value  # 假设要读取的列为A列

        # location = location.replace(location_rep,"")
        if location is not None and url is not None:
            if location in url:
                included_worksheet.append([cell.value for cell in source_worksheet[row]])
            else:
                excluded_worksheet.append([cell.value for cell in source_worksheet[row]])
        else:
            error_worksheet.append([cell.value for cell in source_worksheet[row]])

    # 保存新的工作簿到不同的文件
    included_workbook.save('../data/目标位置5正确记录.xlsx')
    excluded_workbook.save('../data/目标位置5不正确记录.xlsx')
    error_workbook.save('../data/目标位置5错误记录.xlsx')

    # 保存修改后的Excel文件
    workbook.close()


if __name__ == '__main__':
    run()
    # url = "大叔大婶sasda中国上海市浦东新区航春路51号啊实打实的"  # 假设要读取的列为A列
    # location = "上海市浦东新区航春路51号"  # 假设要读取的列为A列
    # if location is not None and url is not None:
    #     if location in url:
    #         print("true")
    #     else:
    #         print("false")
    #
    # url = "大叔大婶sasda中国上海市浦东新区航春路啊实打实的"  # 假设要读取的列为A列
    # location = "上海市浦东新区航春路51号"  # 假设要读取的列为A列
    # if location is not None and url is not None:
    #     if location in url:
    #         print("true")
    #     else:
    #         print("false")
