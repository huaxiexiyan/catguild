import json

import requests
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook


def get_location(locations):
    r"""
    将地址换化为坐标
    :return:
    """
    # 定义URL和参数
    url = "https://restapi.amap.com/v3/assistant/coordinate/convert"
    params = {"key": "", "locations": locations, "coordsys": "gps", "output": "JSON"}
    # 发送GET请求
    return requests.get(url, params=params).json()


def get_location_map(location):
    r"""
    将地址换化为坐标
    :return:
    """
    # 定义URL和参数
    url = "https://restapi.amap.com/v3/geocode/regeo"
    params = {"key": "", "location": location, "extensions": "all", 'radius': '500',
              'poitype': '141200|141201|141202|141203|141204|141207',"output": "JSON"}
    # 发送GET请求
    return requests.get(url, params=params).json()


def run():
    workbook = load_workbook(filename='../data/目标位置正确记录.xlsx')
    source_worksheet = workbook['位置']

    for row in range(2, source_worksheet.max_row + 1):
        # 读取某一列的值
        location_l = source_worksheet['G' + str(row)].value  # 假设要读取的列为A列
        location_r = source_worksheet['F' + str(row)].value  # 假设要读取的列为A列
        location = location_l + "," + location_r
        location_json = get_location(location)

        source_worksheet['K' + str(row)].value = json.dumps(location_json)
        source_worksheet['L' + str(row)].value = location_json['locations']
        # source_worksheet['M' + str(row)].value = location_url
        # source_worksheet['N' + str(row)].value = location_url

    # 保存新的工作簿到不同的文件
    # 141200|141201|141202|141203|141204|141205|141206|141207
    workbook.save('../data/目标位置正确记录位置解析.xlsx')

    # 保存修改后的Excel文件
    workbook.close()

def run2():
    workbook = load_workbook(filename='../data/目标位置正确记录位置解析.xlsx')
    source_worksheet = workbook['位置']

    index  = 1
    for row in range(2, source_worksheet.max_row + 1):
        try:
            # 读取某一列的值
            print(f"{index}: "+ source_worksheet['D' + str(row)].value)
            index = index + 1
            location = source_worksheet['L' + str(row)].value  # 假设要读取的列为A列
            location_json = get_location_map(location)
            source_worksheet['M' + str(row)].value = json.dumps(location_json)
            location_json_str = ''
            location_json_str = location_json_str + location_json['regeocode']['formatted_address'] + ';'
            pois = location_json["regeocode"]["pois"]
            for poi in pois:
                location_json_str = location_json_str + poi["name"] + "@" + poi["address"] + "@" + poi["poiweight"] + ";"
            source_worksheet['N' + str(row)].value = location_json_str
        except Exception as e:
            print("Error parsing JSON or extracting data from:")

    # 保存新的工作簿到不同的文件
    # 141200|141201|141202|141203|141204|141205|141206|141207
    workbook.save('../data/目标位置正确记录位置反解析.xlsx')

    # 保存修改后的Excel文件
    workbook.close()

def run3():
    workbook = load_workbook(filename='../data/目标位置正确记录位置反解析.xlsx')
    source_worksheet = workbook['位置']

    index  = 1
    for row in range(2, source_worksheet.max_row + 1):
        try:
            # 读取某一列的值
            print(f"{index}: "+ source_worksheet['D' + str(row)].value)
            index = index + 1
            location = source_worksheet['D' + str(row)].value  # 假设要读取的列为A列
            location_result = source_worksheet['N' + str(row)].value
            if location in location_result:
                source_worksheet['O' + str(row)].value = '基本一致'
            else:
                source_worksheet['O' + str(row)].value = '肉眼进一步判定'
        except Exception as e:
            print("Error parsing JSON or extracting data from:")

    # 保存新的工作簿到不同的文件
    # 141200|141201|141202|141203|141204|141205|141206|141207
    workbook.save('../data/目标位置正确记录位置反解析判定.xlsx')

    # 保存修改后的Excel文件
    workbook.close()

if __name__ == '__main__':
    # run()
    # run3()
    print(get_location_map('121.693062065973,31.171753472223'))
