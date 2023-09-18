import logging

import requests
import googlemaps
from openpyxl.reader.excel import load_workbook


def get_google_map_api(address):
    r"""
    将地址换化为坐标
    :return:
    """
    # 定义URL和参数
    url = "https://maps.googleapis.com/maps/api/geocode/json"
    params = {"key": "", "address": address, "city": "310115", "output": "JSON"}
    # 发送GET请求
    return requests.get(url, params=params).json()


def get_location(location):
    # 请求api
    result_json = get_google_map_api(location)
    gmaps = googlemaps.Client(key='')
    # Geocoding an address
    geocode_result = gmaps.geocode('1600 Amphitheatre Parkway, Mountain View, CA')
    # 解析经纬度
    try:
        return 0, 0, ""
    except Exception:
        logging.exception("Error message")
        return 0, 0, ""


def load_excel():
    workbook = load_workbook(
        filename='../data/目标位置.xlsx')
    worksheet = workbook['位置']
    for row in range(2, worksheet.max_row + 1):
        try:
            # 读取某一列的值
            column_value = worksheet['E' + str(row)].value  # 假设要读取的列为A列
            print("获取学校门牌号【{}】".format(column_value))
            # 根据门牌号查询经纬度
            longitude, latitude, location_url = get_location("中国" + column_value)
            # 根据读取的值填充另一列
            worksheet['F' + str(row)].value = longitude
            worksheet['G' + str(row)].value = latitude
            worksheet['H' + str(row)].value = location_url
        except Exception:
            logging.exception("Error message")

    # 保存修改后的Excel文件
    workbook.save('../data/目标位置2-1.xlsx')
    workbook.close()


def run():
    workbook = load_workbook(filename='../data/目标位置2-1.xlsx')
    worksheet = workbook['位置']
    for row in range(2, worksheet.max_row + 1):
        # 读取某一列的值
        column_value = worksheet['E' + str(row)].value  # 假设要读取的列为A列
        print("获取学校门牌号【{}】".format(column_value))
        # 根据门牌号查询经纬度
        longitude, latitude, original_result = get_location(driver, "中国" + column_value)
        # 根据读取的值填充另一列
        worksheet['F' + str(row)].value = longitude
        worksheet['G' + str(row)].value = latitude
        worksheet['H' + str(row)].value = original_result

    # 保存修改后的Excel文件
    workbook.save('../data/目标位置3.xlsx')
    workbook.close()


if __name__ == '__main__':
    # run()
    gmaps = googlemaps.Client(key='')
    # Geocoding an address
    # def geocode(client, address=None, place_id=None, components=None, bounds=None, region=None,
    #             language=None)
    geocode_result = gmaps.geocode(address='上海市浦东新区栏学路495号',  language='zh-CN')
    print(geocode_result)
    # 121.66994 31.16272
    # def reverse_geocode(client, latlng, result_type=None, location_type=None,
    #                     language=None)
    # reverse_geocode_result = gmaps.reverse_geocode(latlng=(31.16272, 121.66994), language='zh-CN')
    # print(reverse_geocode_result)
