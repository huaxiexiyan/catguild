import logging
import time
from urllib.parse import unquote

from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait


def download_html_selenium(keyword):
    r"""
    下载html文档
    :return:
    """
    try:
        driver = webdriver.Chrome()

        # 打开百度百科首页：使用 WebDriver 实例打开百度百科首页
        driver.get("https://www.google.com/maps/?hl=zh-cn")
        # 定位输入框并输入文字
        input_box = driver.find_element(By.ID, "searchboxinput")
        input_box.send_keys(keyword)
        # 等 3s 出现下拉框
        time.sleep(3)

        # 等待下拉框
        wait = WebDriverWait(driver, 10)
        wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "sW9vGe")))

        css_selector = "div.sW9vGe img[src*='//maps.gstatic.com/consumer/images/icons/2x/place_grey650.png']"
        img_element = driver.find_element(By.CSS_SELECTOR, css_selector)
        img_element.click()

        # 点击进入词条按钮
        # button = driver.find_element(By.ID, "searchbox-searchbutton")
        # button.click()

        # 获取结果网页
        time.sleep(5)

        # 获取重定向后的 URL 地址
        redirected_url = driver.current_url
        # 输出重定向后的 URL 地址
        decoded_url = unquote(redirected_url)
        print("重定向后的 URL 地址：", decoded_url)
        driver.quit()
        return decoded_url
    except Exception:
        logging.exception("Error message")


def download_html_selenium_new(driver, keyword):
    r"""
    下载html文档
    :return:
    """
    input_box = None
    try:
        time.sleep(1)
        # 定位输入框并输入文字
        input_box = driver.find_element(By.ID, "searchboxinput")
        input_box.send_keys(keyword)
        # 等 3s 出现下拉框
        time.sleep(3)

        # 等待下拉框
        wait = WebDriverWait(driver, 10)
        wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "sW9vGe")))

        # css_selector = "div.sW9vGe img[src*='//maps.gstatic.com/consumer/images/icons/2x/place_grey650.png']"
        # img_element = driver.find_element(By.CSS_SELECTOR, css_selector)
        # img_element.click()

        # 选择带数字的
        xpath_expression = "(//div[@class='sW9vGe']//span[contains(text(), '0') or contains(text(), '1') or contains(text(), '2') or contains(text(), '3') or contains(text(), '4') or contains(text(), '5') or contains(text(), '6') or contains(text(), '7') or contains(text(), '8') or contains(text(), '9')])[1]"
        div_element_with_number = driver.find_element(By.XPATH, xpath_expression)
        div_element_with_number.click()

        # 点击进入词条按钮
        # button = driver.find_element(By.ID, "searchbox-searchbutton")
        # button.click()

        # 获取结果网页
        time.sleep(5)

        # 获取重定向后的 URL 地址
        redirected_url = driver.current_url
        # 输出重定向后的 URL 地址
        decoded_url = unquote(redirected_url)
        print("重定向后的 URL 地址：", decoded_url)

        return decoded_url
    except Exception:
        logging.exception("Error message")
        return ""
    finally:
        # 操作完后需要清空选择框
        input_box.clear()


def get_location(driver, location):
    location_url = download_html_selenium_new(driver, location)
    try:
        if location_url != "":
            # 查找 @ 的位置
            at_index = location_url.find("@")

            # 查找 ,17z 的位置
            comma_index = location_url.find(",17z")

            # 提取经度和纬度坐标
            coordinates = location_url[at_index + 1: comma_index]
            longitude, latitude = coordinates.split(",")
            return longitude, latitude, location_url
        else:
            return 0, 0, ""
    except Exception:
        logging.exception("Error message")
        return 0, 0, ""


def run():
    # 打开浏览器
    driver = webdriver.Chrome()
    # 打开百度百科首页：使用 WebDriver 实例打开百度百科首页
    driver.get("https://www.google.com/maps/?hl=zh-cn")

    workbook = load_workbook(filename='../data/目标位置4错误记录.xlsx')
    worksheet = workbook['位置']
    for row in range(2, worksheet.max_row + 1):
        # 读取某一列的值
        column_value = worksheet['E' + str(row)].value  # 假设要读取的列为A列
        column_value_re = worksheet['C' + str(row)].value  # 假设要读取的列为A列
        column_value = column_value.replace(column_value_re,"")
        print("获取学校门牌号【{}】".format(column_value))
        # 根据门牌号查询经纬度
        longitude, latitude, location_url = get_location(driver, "中国" + column_value)
        # 根据读取的值填充另一列
        worksheet['F' + str(row)].value = longitude
        worksheet['G' + str(row)].value = latitude
        worksheet['H' + str(row)].value = location_url

    # 保存修改后的Excel文件
    workbook.save('../data/目标位置5.xlsx')
    workbook.close()
    # 关闭浏览器
    driver.quit()


if __name__ == '__main__':
    run()
