from openpyxl import load_workbook

from simple_spider.pudong_school.spider.gaode_location import get_location, gcj02_to_wgs84


def load_excel():
    workbook = load_workbook(
        filename='../data/目标位置.xlsx')
    worksheet = workbook['位置']
    index = 1
    for row in range(2, worksheet.max_row + 1):
        if index > 1:
            break
        # 读取某一列的值
        column_value = worksheet['E' + str(row)].value  # 假设要读取的列为A列
        print("获取学校门牌号【{}】".format(column_value))
        # 根据门牌号查询经纬度
        gaode_location_json = get_location("上海市浦东新区栏学路495号")
        print("获取学校门牌号经纬度json【{}】".format(gaode_location_json))
        if gaode_location_json is not None:
            geocodes_json = gaode_location_json['geocodes'][0]
            formatted_address = geocodes_json['formatted_address']
            location_level = geocodes_json['level']
            location = geocodes_json['location']
            location_split = location.split(",")
            print("获取学校查询地址【{}】".format(formatted_address))
            # 转换坐标系
            gcj_lng = location_split[0]
            gcj_lat = location_split[1]
            wgs84_lng, wgs84_lat = gcj02_to_wgs84(gcj_lng, gcj_lat)
            print(wgs84_lng, wgs84_lat)
            # 根据读取的值填充另一列
            worksheet['F' + str(row)].value = gcj_lng
            worksheet['G' + str(row)].value = gcj_lat
            worksheet['H' + str(row)].value = wgs84_lng
            worksheet['I' + str(row)].value = wgs84_lat
            worksheet['J' + str(row)].value = formatted_address
            worksheet['K' + str(row)].value = location_level
        index += 1

    # 保存修改后的Excel文件
    workbook.save('../data/目标位置2-1.xlsx')
    workbook.close()



if __name__ == '__main__':
    # 1、从表格中加载出目标位置
    load_excel()
    # 2、调用api查询经纬度坐标
    # 3、转换为大地国际标准坐标
    # 4、写入表格中
    print("执行结束")
