import hashlib
import logging
import time
from urllib.parse import quote_plus, unquote

import pymongo
import redis
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


def get_proxy():
    r"""
    从代理ip池中随机获取ip
    :return:
    """
    headers = {
        'Authorization': 'Basic '
    }
    return requests.get("https://proxy-pool.catguild.cn/get/", headers=headers).json()


def delete_proxy(proxy):
    r"""
    删除代理
    :param proxy:
    :return:
    """
    headers = {
        'Authorization': 'Basic '
    }
    requests.get("https://proxy-pool.catguild.cn/delete/?proxy={}".format(proxy), headers=headers)


def download_html_proxy(keyword):
    r"""
    下载html文档
    :return:
    """
    retry_count = 100
    proxy = get_proxy().get("proxy")
    while retry_count > 0:
        try:
            print("成功获取代理url: ", proxy)
            # 使用代理访问
            html = requests.get("http://baike.baidu.com/item/{}".format(keyword),
                                proxies={"http": "http://{}".format(proxy)},
                                timeout=30)
            if html.status_code == 200:
                if "百度百科-验证" in html.text or "百度百科" not in html.text:
                    print("代理url不可用,将删除: ", proxy)
                    retry_count -= 1
                    delete_proxy(proxy)
                    proxy = get_proxy().get("proxy")
                    time.sleep(30)
                    continue
            return html
        except Exception as e:
            logging.exception("Error message")
            retry_count -= 1
            # 删除代理池中代理
            print("代理url不可用,将删除: ", proxy)
            delete_proxy(proxy)
            time.sleep(30)
            continue
    return None


def download_html(keyword):
    r"""
    下载html文档
    :return:
    """
    try:
        # 使用代理访问
        time.sleep(1)
        html = requests.get("https://baike.baidu.com/item/{}".format(keyword), timeout=5)
        if html.status_code == 200:
            soup = BeautifulSoup(html.text, 'lxml')
            if "百度百科-验证" not in soup.title.string and "百度百科" in soup.title.string:
                return html
        return None
    except Exception as e:
        logging.exception("Error message", e)


def download_html_selenium(driver, collection, keyword):
    r"""
    下载html文档
    :return:
    """
    try:
        # 打开百度百科首页：使用 WebDriver 实例打开百度百科首页
        driver.get("https://baike.baidu.com/")
        # 定位输入框并输入文字
        input_box = driver.find_element(By.ID, "query")
        input_box.send_keys(keyword)
        # 点击输入框来触发下拉框的出现
        input_box.click()
        # 等待下拉框出现并选择第一个候选项
        wait = WebDriverWait(driver, 10)
        dropdown = wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "sug-lemma_item")))
        dropdown.click()
        # 点击进入词条按钮
        button = driver.find_element(By.ID, "search")
        button.click()
        # 获取结果网页
        wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "content")))
        result_page_url = driver.current_url
        html = driver.page_source
        if html is not None:
            store_html(md5_hash, collection, keyword, html, result_page_url)
            print("成功下载【{}】".format(keyword))
        return driver.page_source
    except Exception:
        logging.exception("Error message")


def store_html(md5_hash, collection, school_name, html, url):
    md5_hash.update(school_name.encode('utf-8'))
    data = {"id": md5_hash.hexdigest(),
            "content_url": url,
            "excel_school_name": school_name,
            "baike_school_name": unquote(url).split("/item/")[1].split("?")[0],
            "content_html": html}
    collection.insert_one(data)


def load_target_school():
    workbook = load_workbook(
        filename='/simple_spider/pudong_school/data/学校情况表.xlsx')
    worksheet = workbook['学校']
    # 选择要读取的列号（假设是第一列）
    column_number = 3
    # 遍历指定列的所有单元格
    column_values = []
    for row in worksheet.iter_rows(min_row=1, min_col=column_number, values_only=True):
        column_values.append(row[0])
    column_values = list(set(column_values))
    # 打印列的所有值
    for value in column_values:
        print(value)
    load_redis(column_values)
    workbook.close()


def load_redis(data):
    r = get_redis_client()
    # 设置键值对
    for item in data:
        r.rpush('school_name', item)
    # 获取键的值
    value = r.lrange('school_name', 0, -1)
    print(value)


def get_redis_client():
    return redis.Redis(host='', password="", port=16379, db=1)


def get_mongo_client():
    client = pymongo.MongoClient(
        "mongodb://%s:%s@%s" % (quote_plus(""), quote_plus(""), ""))
    return client["catguild_prod"]


def get_target_school_name():
    r = get_redis_client()
    school_name_list = r.lrange('school_name', 0, -1)
    data = []
    for item in school_name_list:
        data.append(item.decode('utf-8'))
    return data


if __name__ == '__main__':
    school_name_list = get_target_school_name()
    print(school_name_list)
    driver = webdriver.Chrome()
    md5_hash = hashlib.md5()
    db = get_mongo_client()
    collection = db["school"]
    list_size = len(school_name_list)
    print("总记录数【{}】".format(list_size))
    for school_name in school_name_list:
        print("准备下载html【{}】".format(school_name))
        html = download_html_selenium(driver, collection, school_name)
        time.sleep(30)
        print("暂停 30 s 后继续")
        list_size = list_size - 1
        print("还剩记录数【{}】".format(list_size))
    driver.quit()
