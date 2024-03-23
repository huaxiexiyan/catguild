from openpyxl import load_workbook


def load_excel_and_save():
    workbook_source = load_workbook(filename='./data/资源单号详情.xlsx', read_only=True)
    original_ws = workbook_source['总表']

    # 创建一个新的Excel文件
    workbook_result = load_workbook(filename='./data/单号提取结果.xlsx')
    result_ws = workbook_result.active
    result_ws.delete_rows(1, result_ws.max_row)


    # 加载关键字
    list_keyword = load_excel_target()
    # 创建一个集合来存储匹配到的关键词
    matched_keywords = set()

    # 遍历原始Excel文件中的每一行
    print("========= 开始提取目标单号 =========")
    # 保留表头
    first_row = next(original_ws.iter_rows(values_only=True))
    result_ws.append(first_row)
    for row in original_ws.iter_rows(min_row=2, values_only=True):
        # 检查当前行的A列是否包含关键词
        if any(keyword in row[0] for keyword in list_keyword):  # 这里假设关键词在A列中
            # 如果匹配到关键词，则将整行复制到新的Excel文件中
            print("目标单号【{}】是目标单号".format(row[0]))
            result_ws.append(row)
            matched_keywords.add(row[0])
        else:
            print("目标单号【{}】不是目标单号".format(row[0]))
    print("========= 提取目标单号完成 =========")
    # 找出未匹配到的关键词
    unmatched_keywords = set(list_keyword) - matched_keywords
    # 输出未匹配到的关键词
    print("未匹配到的单号：", unmatched_keywords)

    # 保存修改后的Excel文件
    workbook_source.close()
    workbook_result.save(filename='./data/单号提取结果.xlsx')


def load_excel_target():
    print("========= 开始加载目标单号 =========")
    workbook_target = load_workbook(filename='./data/目标提取单号.xlsx', read_only=True)
    target_ws = workbook_target['单号']

    # 遍历原始Excel文件中的每一行
    list_keyword = []
    for row in target_ws.iter_rows(min_row=2, values_only=True):
        list_keyword.append(row[1])

    print("单号总数【{}】".format(len(list_keyword)))
    unique_list = list(set(list_keyword))
    print("去重后单号总数【{}】".format(len(unique_list)))
    # 关闭流
    workbook_target.close()
    print("========= 加载目标单号完成 =========")
    return list_keyword


if __name__ == '__main__':
    """
    1、读取位置信息
    2、解析出excel
    """
    load_excel_and_save()
    # load_excel_target()
